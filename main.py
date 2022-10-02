import openpyxl
book = openpyxl.load_workbook("Task 1.xlsx", read_only=True)
sheet = book.worksheets[1]

variants = list()
for row in range(3, sheet.max_row + 1):
    variants.append(sheet[row][0].value)

t1y1 = list()
for row in range(3, sheet.max_row + 1):
    t1y1.append(sheet[row][1].value)

t1m1 = list()
for row in range(3, sheet.max_row + 1):
    t1m1.append(sheet[row][2].value)

t1y2 = list()
for row in range(3, sheet.max_row + 1):
    t1y2.append(sheet[row][3].value)

t1m2 = list()
for row in range(3, sheet.max_row + 1):
    t1m2.append(sheet[row][4].value)

t1v1 = list()
for row in range(3, sheet.max_row + 1):
    t1v1.append(sheet[row][5].value)

t1Tmoor = list()
for row in range(3, sheet.max_row + 1):
    t1Tmoor.append(sheet[row][6].value)

# Delta T
dt = list()
for i in range(len(variants)):
    dtr = (t1y2[i] - t1y1[i]) * 12 + t1m2[i] - t1m1[i]
    dt.append(dtr)

# V2 memory
t1v2 = list()
for i in range(len(variants)):
    t1v2r = t1v1[i] * 2**(dt[i]/t1Tmoor[i])
    t1v2r = round(t1v2r,2)
    t1v2.append(t1v2r)


print(t1v2)