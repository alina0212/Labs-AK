import openpyxl
book = openpyxl.load_workbook("date.xlsx", read_only=True)
sheet = book.worksheets[0]

variants = list()
for row in range(3, sheet.max_row + 1):
    variants.append(sheet[row][0].value)

t2y1 = list()
for row in range(3, sheet.max_row + 1):
    t2y1.append(sheet[row][1].value)

t2m1 = list()
for row in range(3, sheet.max_row + 1):
    t2m1.append(sheet[row][2].value)

t2y2 = list()
for row in range(3, sheet.max_row + 1):
    t2y2.append(sheet[row][3].value)

t2m2 = list()
for row in range(3, sheet.max_row + 1):
    t2m2.append(sheet[row][4].value)

t2v1 = list()
for row in range(3, sheet.max_row + 1):
    t2v1.append(sheet[row][5].value)

t2Tmoor = list()
for row in range(3, sheet.max_row + 1):
    t2Tmoor.append(sheet[row][6].value)
print(t2m1)
